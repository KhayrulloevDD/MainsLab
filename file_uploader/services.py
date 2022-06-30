import openpyxl
import random
from file_uploader.models import Organization, Bill
from file_uploader.serializers import OrganizationSerializer, BillSerializer


def parse_client_org_xlsx_file(client_org_file):
    work_book = openpyxl.load_workbook(client_org_file)

    clients = []
    sheet_client = work_book['client']
    for client_name in range(2, sheet_client.max_row + 1):
        clients.append({
            "name": sheet_client[client_name][0].value,
        })

    organizations = []
    sheet_organization = work_book['organization']
    for organization in range(2, sheet_organization.max_row + 1):
        if sheet_organization[organization][0].value:
            organizations.append({
                "client_name": sheet_organization[organization][0].value,
                "name": sheet_organization[organization][1].value,
                "address": f"Адресс: {sheet_organization[organization][2].value}",
            })

    return {
        "clients": clients,
        "organizations": organizations
    }


def parse_bills_xlsx_file(bills_file):
    work_book = openpyxl.load_workbook(bills_file)

    bills = []
    sheet_bills = work_book['Лист1']
    for bill in range(2, sheet_bills.max_row + 1):
        if sheet_bills[bill][0].value:
            service_class, service_name = service_classifier(sheet_bills[bill][5].value)
            fraud_score = fraud_detector(sheet_bills[bill][5].value)
            bills.append({
                "client_name": sheet_bills[bill][0].value,
                "client_org": sheet_bills[bill][1].value,
                "number": sheet_bills[bill][2].value,
                "summary": sheet_bills[bill][3].value,
                "date": sheet_bills[bill][4].value.date(),
                "service": sheet_bills[bill][5].value,
                "fraud_score": fraud_score,
                "service_class": service_class,
                "service_name": service_name,
            })

    return {
        "bills": bills
    }


def replace_client_name_to_client_id(organizations, clients):
    for organization in organizations:
        for client in clients:
            if organization['client_name'] == client['name']:
                organization['client_name'] = client['id']
                break
    return organizations


def replace_client_name_and_client_org_to_ids(bills, clients, organizations):
    for bill in bills:
        for client in clients:
            if bill['client_name'] == client['name']:
                bill['client_name'] = client['id']
                break
        for organization in organizations:
            if bill['client_org'] == organization['name']:
                bill['client_org'] = organization['id']
                break
    return bills


def fraud_detector(string: str):
    return round(random.random(), 2)


def service_classifier(string: str):
    choices = {
        1: "консультация",
        2: "лечение",
        3: "стационар",
        4: "диагностика",
        5: "лаборатория’"
    }
    return random.choice(list(choices.items()))


def count_fraud_weight():
    organizations = Organization.objects.all()
    for organization in organizations:
        fraud_weight = Bill.objects.filter(client_org=organization.id, fraud_score__gte=0.9).count()
        organization.fraud_weight = fraud_weight
        organization.save()
